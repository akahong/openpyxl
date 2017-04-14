#encoding:utf-8

"""
 官方文档：http://packages.python.org/openpyxl/api.html
"""
from openpyxl import Workbook
import datetime

#workbook
#There is no need to create a file on the filesystem to get started with openpyxl.
# Just import the Workbook class and start using it
#创建一个对象
wb=Workbook()


#worksheet
#workbook至少创建一个worksheet
#通过openpyxl.workbook.Workbook.active()得到worksheet
ws=wb.active

#该方法使用_active_sheet_index属性, 默认会设置0，也就是第一个worksheet。除非手动修改，否则使用active方法得到都是第一个worksheet。
#你也可以创建worksheets，通过 openpyxl.workbook.Workbook.create_sheet() 方法：
ws1=wb.create_sheet('mysheet') #插入到最后（default）
ws1=wb.create_sheet('mysheet',0) #插入到最开始的位置

#创建的sheet的名称会自动创建，按照sheet，sheet1，sheet2自动增长，通过title属性可以修改其名称。
ws.title = "New Title"

#默认的sheet的tab是白色的，可以通过 RRGGBB颜色来修改sheet_properties.tabColor属性从而修改sheet tab按钮的颜色:
ws.sheet_properties.tabColor = "1072BA"

#当你设置了sheet的名称，可以将其看成workbook中的一个key。也可以使用openpyxl.workbook.Workbook.get_sheet_by_name() 方法
ws3 = wb["New Title"]
ws4 = wb.get_sheet_by_name("New Title")
ws is ws3 is ws4
#返回结果为True


#查看workbook中的所有worksheets名称:openpyxl.workbook.Workbook.get_sheet_names()
print wb.sheetnames
#返回所有的sheet name列表

# 显示有多少张表
print   "Worksheet range(s):", wb.get_named_ranges()
print   "Worksheet name(s):", wb.get_sheet_names()

# 取第一张表
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])

# 显示表名，表行数，表列数
print   "Work Sheet Titile:", ws.title
print   "Work Sheet Rows:", ws.get_highest_row()
print   "Work Sheet Cols:", ws.get_highest_column()


#操作数据

#访问单元格
#单元格可以看作是worksheet的key，通过key去访问单元格中的数据
c=ws['A4'] #指定单元格的值，直接赋值
#直接返回A4单元格，如果不存在则会自动创建一个。

d=ws.cell(row=4,column=2,value=10) #使用openpyxl.worksheet.Worksheet.cell()方法操作某行某列的某个值

#当worksheet在内存中被创建时,是没有包含cells的，cells是在首次访问时创建.
#可以循环在内存中创建cells，这时不指定他们的值也会创建该cells些：(创建100x100cells)
for i in range(1,101):
    for j in range(1,101):
        ws.cell(row=i,column=j)

#访问许多cell

#通过切片ranges指定许多cells
cell_range=ws['A1':'C1']
#同样也可以Ranges rows 或者columns
colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

#也可以使用 openpyxl.worksheet.Worksheet.iter_rows() 方法:(需要指定行->行，截止列)
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

#也可以使用 openpyxl.worksheet.Worksheet.iter_cols() 方法:(需要指定列->列，截止行)
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
   for cell in col:
       print(cell)

#如果你需要遍历所有文件的行或列,可以使用openpyxl.worksheet.Worksheet.rows 属性
tuple(ws.rows)
#或者 openpyxl.worksheet.Worksheet.columns 属性
tuple(ws.columns)


#saving to a file
#最简单最安全的方法保存workbook是使用openpyxl.workbook.Workbook对象的 openpyxl.workbook.Workbook.save()方法
wb = Workbook()
wb.save('balances.xlsx')
#保存的默认位置在python的根目录下。
#注意：会自动覆盖已经存在文件名的文件。


#loading from a file
#像写一样我们可以导入openpyxl.load_workbook()已经存在的workbook
from openpyxl import load_workbook
wb2 = load_workbook('test.xlsx')
print wb2.get_sheet_names()

