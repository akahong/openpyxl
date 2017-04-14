#encoding:utf-8

from openpyxl import Workbook
import datetime

wb=Workbook() #创建一个表对象
ws=wb.active # 创建一个sheet对象
ws["A1"]=41 #将数据直接分配到单元格中
ws.append([1,2,3]) #附加一行，从第一列开始附加
ws['A3']=datetime.datetime.now().strftime('%Y-%m-%d') #python 类型会被自动转换
wb.save('sample.xlsx')