# -*- coding: utf-8 -*-
"""
http://openpyxl.readthedocs.io/en/default/usage.html
"""

# workbook相关
from openpyxl import Workbook
from openpyxl.compat import range
#数字转为列字母的方法
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
#一种格式化字符串的函数str.format()
#如：'{0},{1}'.format('kzc',18)
#Out[1]: 'kzc,18'
#In [2]: '{},{}'.format('kzc',18)
#Out[2]: 'kzc,18'
#In [3]: '{1},{0},{1}'.format('kzc',18)
#Out[3]: '18,kzc,18
#get_column_letter（）数字转为列字母的方法
print(ws3['AA10'].value)
wb.save(filename=dest_filename)