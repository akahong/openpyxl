# -*- coding: utf-8 -*-

from openpyxl.reader.excel import load_workbook
import json

# 读取excel2007文件
wb = load_workbook(filename=r'test_book.xlsx')

# 显示有多少张表
print "Worksheet range(s):", wb.get_named_ranges()
print "Worksheet name(s):", wb.get_sheet_names()

# 取第一张表
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])
ws=wb.active

# 显示表名，表行数，表列数
print "Work Sheet Titile:", ws.title
print "Work Sheet Rows:", ws.max_row
print "Work Sheet Cols:", ws.max_column


# 建立存储数据的字典
data_dic = {}

# 把数据存到字典中
for rx in range(1, ws.max_row + 1):
    temp_list = []
    pid = rx
    w1 = ws.cell(row=rx, column=1).value
    w2 = ws.cell(row=rx, column=2).value
    w3 = ws.cell(row=rx, column=3).value
    w4 = ws.cell(row=rx, column=4).value
    temp_list = [w1, w2, w3, w4]

    data_dic[pid] = temp_list

# 打印字典数据个数
print 'Total:%d' % len(data_dic)
print json.dumps(data_dic, encoding="UTF-8", ensure_ascii=False)
